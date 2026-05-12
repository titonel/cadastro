"""
Producao SIRESP — parser de arquivo XLS/XLSX exportado do portal SIRESP.

Estrutura esperada:
  B2  : "Período:DD-MM-AAAAaDD-MM-AAAA"
  Linha 9  : cabeçalhos (ignorados — usamos COLUNAS_SIRESP)
  Linha 10+: dados

Regras de identificação na coluna A:
  - Nome da agenda  → Title Case (ex: "Oftalmologia - Catarata")
  - Nome do médico  → UPPER CASE  (ex: "JOAO DA SILVA")
  - "Total Geral"   → linha de rodapé — encerra o processamento
  - Demais linhas de total/subtotal/rodapé → ignoradas

Dependencia: openpyxl (suporta .xlsx e .xlsm; para .xls legado use xlrd<=1.2.0)
"""

import io
import re
from datetime import date, datetime
from typing import Optional

import openpyxl

from .models import (
    UploadProducao, ProducaoAgenda, ProducaoMedico,
    StatusImportacao, COLUNAS_SIRESP,
)

# Lista canônica de agendas para validação cruzada (opcional)
AGENDAS_CONHECIDAS = {
    "Anestesiologia - Avaliação Pré-cirúrgica",
    "Cardiologia",
    "Cardiologia - Hipertensão",
    "Cardiologia - Saude do Homem",
    "Cirurgia Geral",
    "Cirurgia Geral - Pós Operatório",
    "Cirurgia Pediátrica",
    "Cirurgia Pediátrica - Pós Operatório",
    "Cirurgia Plástica",
    "Cirurgia Plástica - Pós Operatório",
    "Cirurgia Plástica - Tumor Pele e Outros",
    "Cirurgia Vascular",
    "Cirurgia Vascular - Pós Operatório",
    "Coloproctologia",
    "Coloproctologia - Pós Operatório",
    "Dermatologia",
    "Dermatologia - Pós Cirúrgica",
    "Endocrinologia",
    "Enfermagem",
    "Enfermagem - Biópsia de Próstata",
    "Enfermagem - Cardiologia",
    "Enfermagem - Cistoscopia",
    "Enfermagem - Hipertensão",
    "Enfermagem - Microcefalia - Linha de Cuidado",
    "Enfermagem - Multiorientações",
    "Enfermagem - Orientação Cirúrgica",
    "Enfermagem - Saude do Homem",
    "Farmacêutico - Orientação de Protocolo",
    "Farmácia",
    "Gastroclínica",
    "Gastroclínica - Triagem Colonoscopia",
    "Gastroclínica - Triagem Endoscopia",
    "Mastologia",
    "Mastologia - Pós-operatório",
    "Neurologia",
    "Neurologia - Sífilis Congênita",
    "Neurologia Pediátrica",
    "Nutrição",
    "Oftalmologia",
    "Oftalmologia - Avaliação Cirúrgica",
    "Oftalmologia - Avaliação Pós-cirúrgica",
    "Oftalmologia - Catarata",
    "Oftalmologia - Microcefalia - Linha de Cuidado",
    "Oftalmologia - Pterígio",
    "Oftalmologia - Reflexo Vermelho",
    "Oftalmologia - Retina",
    "Oftalmologia - Sífilis Congênita",
    "Ortopedia",
    "Ortopedia - Pós-operatório",
    "Otorrinolaringologia",
    "Otorrinolaringologia - Nasofibroscopia",
    "Otorrinolaringologia - Sífilis Congênita",
    "Pneumologia",
    "Pneumologia Pediátrica",
    "Serviço Social",
    "Urologia",
    "Urologia - Avaliação Pós-cirúrgica",
    "Urologia - Saude do Homem",
    "Urologia - Vasectomia",
}

# Padrão do campo de período na célula B2
_PERIODO_RE = re.compile(
    r"Per[ií]odo[:\s]*"
    r"(\d{2}-\d{2}-\d{4})"
    r"[aA]"
    r"(\d{2}-\d{2}-\d{4})",
    re.IGNORECASE,
)


def _parse_date(s: str) -> Optional[date]:
    """Converte 'DD-MM-AAAA' em date."""
    try:
        return datetime.strptime(s.strip(), "%d-%m-%Y").date()
    except (ValueError, AttributeError):
        return None


def _cell_str(cell) -> str:
    """Retorna o valor de uma célula openpyxl como string limpa."""
    val = cell.value
    if val is None:
        return ""
    return str(val).strip()


def _safe_int(value) -> int:
    try:
        return int(float(str(value).replace(",", ".").strip()))
    except (ValueError, TypeError):
        return 0


def _safe_float(value) -> float:
    try:
        return float(str(value).replace(",", ".").replace("%", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _eh_total_geral(texto: str) -> bool:
    """Retorna True se a linha é o rodapé 'Total Geral' — encerra o processamento."""
    return texto.strip().lower() == "total geral"


def _eh_agenda(texto: str) -> bool:
    """
    Retorna True se o texto corresponde ao nome de uma agenda:
    - Começa com letra maiúscula
    - Não está inteiramente em maiúsculas (não é médico)
    - Tem ao menos 3 caracteres
    - Não é linha de total/subtotal/rodapé
    """
    if not texto or len(texto) < 3:
        return False
    texto = texto.strip()
    palavras_descarte = (
        "total", "subtotal", "grand total", "período",
        "especialidade", "vagas", "agendamentos",
    )
    lower = texto.lower()
    for p in palavras_descarte:
        if lower.startswith(p):
            return False
    letras = [c for c in texto if c.isalpha()]
    if not letras:
        return False
    if all(c.isupper() for c in letras):
        return False  # é médico
    return letras[0].isupper()


def _eh_medico(texto: str) -> bool:
    """Retorna True se todas as letras são maiúsculas (nome de médico)."""
    if not texto or len(texto) < 3:
        return False
    texto = texto.strip()
    letras = [c for c in texto if c.isalpha()]
    if not letras:
        return False
    return all(c.isupper() for c in letras)


def _extrair_linha(row, n_colunas: int) -> dict:
    """
    Extrai os valores de uma linha openpyxl (tupla de células)
    e mapeia aos nomes de campo definidos em COLUNAS_SIRESP.
    """
    campos = {}
    for col_idx in range(min(n_colunas, len(row))):
        nome = COLUNAS_SIRESP[col_idx]
        val = row[col_idx].value
        if val is None:
            val = ""
        elif isinstance(val, str):
            val = val.strip()
        campos[nome] = val
    return campos


def _preencher_campos_numericos(obj, dados: dict):
    """Atribui os campos numéricos a uma instância de ProducaoAgenda ou ProducaoMedico."""
    campos_int = [
        "vagas_ofertadas", "agend_totais", "agend_bolsao", "nao_distribuidas",
        "cota", "extra", "total_geral", "presencial", "teleconsulta",
        "agend_totais_2", "recepcao_ausente", "recepcao_dispensado",
        "recepcao_desistente", "recepcao_nao_informado", "alta",
    ]
    campos_float = [
        "agend_totais_pct", "agend_bolsao_pct", "nao_distribuidas_pct",
        "cota_pct", "extra_pct", "presencial_pct", "teleconsulta_pct",
        "agend_totais_2_pct", "recepcao_ausente_pct", "recepcao_dispensado_pct",
        "recepcao_desistente_pct", "recepcao_nao_informado_pct", "alta_pct",
    ]
    for campo in campos_int:
        if campo in dados:
            setattr(obj, campo, _safe_int(dados[campo]))
    for campo in campos_float:
        if campo in dados:
            setattr(obj, campo, _safe_float(dados[campo]))


def processar_upload(upload_id: int) -> None:
    """
    Lê o XLSX associado ao UploadProducao e persiste
    ProducaoAgenda + ProducaoMedico no banco.

    Levanta exceções — o chamador deve capturar e registrar
    o erro em upload.erro_processamento.
    """
    upload = UploadProducao.objects.get(pk=upload_id)

    # Abre o workbook com openpyxl (data_only=True lê valores calculados, não fórmulas)
    conteudo = upload.arquivo.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(conteudo), data_only=True)
    sheet = wb.active

    # — Extrai período da célula B2 —
    celula_b2 = _cell_str(sheet["B2"])
    match = _PERIODO_RE.search(celula_b2)
    if match:
        upload.data_inicio_periodo = _parse_date(match.group(1))
        upload.data_fim_periodo = _parse_date(match.group(2))

    # — Limpa dados anteriores deste upload —
    upload.agendas.all().delete()

    # — Percorre as linhas a partir da linha 10 (min_row=10) —
    # iter_rows devolve tuplas de células; values_only=False para acessar .value
    agenda_obj: Optional[ProducaoAgenda] = None
    total_agendas = 0
    total_medicos = 0
    n_colunas = len(COLUNAS_SIRESP)

    for row in sheet.iter_rows(min_row=10, max_col=n_colunas):
        col_a = _cell_str(row[0])
        if not col_a:
            continue

        # Linha de rodapé "Total Geral" — encerra o processamento
        if _eh_total_geral(col_a):
            break

        dados = _extrair_linha(row, n_colunas)

        if _eh_agenda(col_a):
            agenda_obj, _ = ProducaoAgenda.objects.get_or_create(
                upload=upload,
                nome_agenda=col_a,
            )
            _preencher_campos_numericos(agenda_obj, dados)
            agenda_obj.save()
            total_agendas += 1

        elif _eh_medico(col_a) and agenda_obj is not None:
            medico_obj = ProducaoMedico(
                agenda=agenda_obj,
                nome_medico=col_a,
            )
            _preencher_campos_numericos(medico_obj, dados)
            medico_obj.save()
            total_medicos += 1

    upload.total_agendas = total_agendas
    upload.total_medicos = total_medicos
    upload.status = StatusImportacao.CONFIRMADO
    upload.erro_processamento = ""
    upload.save()
